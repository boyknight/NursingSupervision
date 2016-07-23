# -*- encoding: utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
import os
import json
import codecs


if __name__ is "__main__":
    doc_path = "/Users/steven/workspace/git/NursingSupervision/Docs"

    for nodes in os.walk(doc_path):
        for file in nodes[2]:
            file_full_path = os.path.join(doc_path, file)

            print(file_full_path)

            # 检查文件扩展名
            ext = file.split(".")[1]
            if ext != "xlsx":
                continue

            # 打开excel文件
            wb = load_workbook(filename=file_full_path)

            first_sheet_name = wb.get_sheet_names()[0]

            sheet = wb.get_sheet_by_name(first_sheet_name)

            max_row = sheet.max_row
            max_column = sheet.max_column

            # 读取科室 (第2行)
            for i in range(1, max_row + 1):
                for j in range(1, max_column + 1):
                    cell = sheet.cell(coordinate=None, row=i, column=j)

                    cell_value = str(cell.value).replace(" ", "")

                    print("cell({}, {}) = {}".format(i, j, cell_value))
