# -*- encoding: utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
import os
import json
import codecs

if __name__ == "__main__":
    doc_path = "/Users/steven/Desktop/1/"

    class_array = list()
    project_array = list()
    issue_array = list()

    class_db_index = 0
    project_db_index = 0
    issue_db_index = 0

    # 搜索所有文件
    for nodes in os.walk(doc_path):
        for file in nodes[2]:
            file_full_path = os.path.join(doc_path, file)
            class_name = file.split(".")[0]
            ext = file.split(".")[1]
            if ext != "xlsx":
                continue

            project_array = list()
            class_dict = {"name": class_name, "projects": project_array}
            class_array.append(class_dict)

            # 打开文件, 读取数据
            wb = load_workbook(filename=file_full_path)

            first_sheet_name = wb.get_sheet_names()[0]

            sheet = wb.get_sheet_by_name(first_sheet_name)
            max_row = sheet.max_row
            max_column = sheet.max_column

            print(file)

            # 确定 项目 和 问题的位置
            project_column_index = 0
            issue_column_index = 0
            score_column_index = 0

            data_row_index = 0

            for i in range(1, max_row + 1):
                for j in range(1, max_column + 1):
                    cell = sheet.cell(coordinate=None, row=i, column=j)

                    cell_value = str(cell.value).replace(" ", "")

                    if cell_value == "项目":
                        project_column_index = j
                        data_row_index = i + 1
                    elif cell_value == "缺陷内容":
                        issue_column_index = j
                        data_row_index = i + 1
                    elif cell_value == "扣分":
                        score_column_index = j
                        data_row_index = i + 1

            # 获取数据
            project_value = None

            for i in range(data_row_index, max_row + 1):
                project_value_temp = sheet.cell(coordinate=None, row=i, column=project_column_index).value
                if project_value_temp is not None:
                    project_value = project_value_temp
                    project_dict = {"name":project_value}
                    project_array.append(project_dict)
                    project_db_index += 1
                    issue_array = list()
                    project_dict["issues"] = issue_array

                issue_value = sheet.cell(coordinate=None, row=i, column=issue_column_index).value
                score_value = sheet.cell(coordinate=None, row=i, column=score_column_index).value

                if issue_value is None:
                    continue

                issue_dict = {"name":issue_value,
                              "score":score_value}

                issue_db_index += 1
                issue_array.append(issue_dict)

        class_db_index += 1

    data_dict = {"classifies":class_array}
    class_array_json = json.dumps(data_dict, skipkeys=False, ensure_ascii=True, check_circular=True,
        allow_nan=True, cls=None, indent=4, separators=None,
        default=None, sort_keys=False)

    total_data = class_array_json

    f = open(os.path.join(doc_path, "init.json"), "tw")
    f.write(total_data)
    f.close()
